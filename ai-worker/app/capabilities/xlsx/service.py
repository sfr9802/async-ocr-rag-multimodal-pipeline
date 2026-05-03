"""XLSX_EXTRACT service and capability wrapper."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from app.capabilities.base import (
    Capability,
    CapabilityError,
    CapabilityInput,
    CapabilityInputArtifact,
    CapabilityOutput,
)
from app.capabilities.xlsx.artifact_builder import (
    XLSX_PIPELINE_VERSION,
    build_output_artifacts,
)

log = logging.getLogger(__name__)

_XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_XLSM_MIME_TYPES = {
    "application/vnd.ms-excel.sheet.macroenabled.12",
}
_XLS_EXTENSIONS = (".xls",)
_XLSX_EXTENSIONS = (".xlsx", ".xlsm")

_SECOND_PASS_MAX_BYTES = 5 * 1024 * 1024
_SECOND_PASS_MAX_CELLS = 50_000
_MAX_COLUMNS = 80
_MAX_ROWS_TO_READ = 10_000
_MAX_CELLS_TO_EXPORT = 10_000
_SECTION_PREVIEW_ROWS = 40
_CHUNK_DATA_ROWS = 50
_MAX_CELL_TEXT = 500
_MAX_UNIT_TEXT = 30_000
_SMALL_TABLE_MAX_CELLS = 2_000
_VOLATILE_FUNCTIONS = ("NOW(", "TODAY(", "RAND(", "RANDBETWEEN(", "OFFSET(", "INDIRECT(")

try:  # openpyxl uses defusedxml automatically when it is installed.
    import defusedxml.ElementTree as _defused_et  # noqa: F401

    _DEFUSEDXML_AVAILABLE = True
except Exception:  # pragma: no cover - depends on environment packaging
    _DEFUSEDXML_AVAILABLE = False


@dataclass
class SheetMetadata:
    name: str
    index: int
    hidden: bool
    max_row: int
    max_column: int
    used_range: str
    tables: list[dict[str, Any]] = field(default_factory=list)
    merged_cells: list[str] = field(default_factory=list)
    formulas: list[dict[str, Any]] = field(default_factory=list)
    hidden_rows: set[int] = field(default_factory=set)
    hidden_columns: set[int] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)


@dataclass
class SheetReadResult:
    rows: list[list[str]]
    row_numbers: list[int]
    cells: list[dict[str, Any]]
    cells_truncated: bool


class XlsxExtractService:
    def __init__(
        self,
        *,
        pipeline_version: str = XLSX_PIPELINE_VERSION,
        include_hidden: bool = False,
    ) -> None:
        self._pipeline_version = pipeline_version
        self._include_hidden = include_hidden

    def run(self, input: CapabilityInput) -> CapabilityOutput:
        artifact = self._pick_input_artifact(input)
        source_record_id = artifact.source_file_id or f"input-artifact:{artifact.artifact_id}"
        mime_type, file_type = self._classify(artifact)
        filename = artifact.filename or f"{artifact.artifact_id}.{file_type}"

        log.info(
            "XLSX extract start jobId=%s artifact=%s fileType=%s",
            input.job_id,
            artifact.artifact_id,
            file_type,
        )
        payload = self._extract(
            artifact.content,
            source_record_id=source_record_id,
            content_type=mime_type,
            filename=filename,
            file_type=file_type,
        )
        return CapabilityOutput(outputs=build_output_artifacts(payload))

    @staticmethod
    def _pick_input_artifact(input: CapabilityInput) -> CapabilityInputArtifact:
        for candidate in input.inputs:
            if candidate.type == "INPUT_FILE":
                return candidate
        if not input.inputs:
            raise CapabilityError("NO_INPUT", "XLSX_EXTRACT job has no input artifacts.")
        raise CapabilityError(
            "UNSUPPORTED_INPUT_TYPE",
            "XLSX_EXTRACT requires an INPUT_FILE artifact; got "
            + ", ".join(sorted({i.type for i in input.inputs})),
        )

    def _classify(self, artifact: CapabilityInputArtifact) -> tuple[Optional[str], str]:
        mime = (artifact.content_type or "").split(";", 1)[0].strip().lower() or None
        filename = (artifact.filename or "").lower()

        if filename.endswith(_XLS_EXTENSIONS):
            raise CapabilityError(
                "UNSUPPORTED_INPUT_TYPE",
                "XLSX_EXTRACT supports .xlsx and read-only .xlsm. Legacy .xls is not supported.",
            )
        if filename.endswith(".xlsm") or mime in _XLSM_MIME_TYPES:
            return mime or "application/vnd.ms-excel.sheet.macroenabled.12", "xlsm"
        if filename.endswith(".xlsx") or mime in _XLSX_MIME_TYPES:
            return mime or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        raise CapabilityError(
            "UNSUPPORTED_INPUT_TYPE",
            "XLSX_EXTRACT supports .xlsx and read-only .xlsm only. Received "
            f"content_type={artifact.content_type!r} filename={artifact.filename!r}.",
        )

    def _extract(
        self,
        content: bytes,
        *,
        source_record_id: str,
        content_type: Optional[str],
        filename: str,
        file_type: str,
    ) -> dict[str, Any]:
        warnings: list[str] = []
        if not _DEFUSEDXML_AVAILABLE:
            warnings.append("defusedxml is not installed; openpyxl XML parsing hardening is unavailable.")
        if file_type == "xlsm":
            warnings.append("XLSM macros are not executed or imported; workbook is read as spreadsheet data only.")

        second_pass = self._collect_second_pass_metadata(content, warnings)
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
        try:
            sheets: list[dict[str, Any]] = []
            visible_sheet_count = 0
            plain_parts: list[str] = []
            for index, worksheet in enumerate(workbook.worksheets):
                sheet_meta = second_pass.get(worksheet.title) or SheetMetadata(
                    name=worksheet.title,
                    index=index,
                    hidden=worksheet.sheet_state != "visible",
                    max_row=worksheet.max_row or 0,
                    max_column=worksheet.max_column or 0,
                    used_range=_used_range(worksheet.max_row or 0, worksheet.max_column or 0),
                )
                sheet_payload = self._extract_sheet(worksheet, sheet_meta)
                sheets.append(sheet_payload)
                if not sheet_payload.get("hidden"):
                    visible_sheet_count += 1
                    compact = (sheet_payload.get("compactText") or "").strip()
                    if compact:
                        plain_parts.append(compact)
        finally:
            workbook.close()

        payload = {
            "fileType": file_type,
            "sourceRecordId": source_record_id,
            "pipelineVersion": self._pipeline_version,
            "extractor": "openpyxl",
            "contentType": content_type,
            "filename": filename,
            "security": {
                "defusedXmlAvailable": _DEFUSEDXML_AVAILABLE,
                "keepLinks": False,
                "macrosExecuted": False,
            },
            "warnings": warnings,
            "workbook": {
                "role": "workbook",
                "sheetCount": len(sheets),
                "visibleSheetCount": visible_sheet_count,
                "sheets": sheets,
            },
            "plainText": "\n\n".join(plain_parts).strip(),
        }
        return payload

    def _collect_second_pass_metadata(
        self,
        content: bytes,
        warnings: list[str],
    ) -> dict[str, SheetMetadata]:
        if len(content) > _SECOND_PASS_MAX_BYTES:
            warnings.append("Workbook skipped metadata second pass because file size exceeds safety limit.")
            return {}
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=False,
                data_only=False,
                keep_links=False,
                keep_vba=False,
            )
        except Exception as ex:
            warnings.append(f"Workbook metadata second pass failed: {type(ex).__name__}: {ex}")
            return {}

        try:
            result: dict[str, SheetMetadata] = {}
            for index, worksheet in enumerate(workbook.worksheets):
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                cell_count = max_row * max_column
                meta = SheetMetadata(
                    name=worksheet.title,
                    index=index,
                    hidden=worksheet.sheet_state != "visible",
                    max_row=max_row,
                    max_column=max_column,
                    used_range=_used_range(max_row, max_column),
                )
                meta.hidden_rows = {
                    row_idx
                    for row_idx, row_dim in worksheet.row_dimensions.items()
                    if row_dim.hidden
                }
                meta.hidden_columns = {
                    _column_index(col_key)
                    for col_key, col_dim in worksheet.column_dimensions.items()
                    if col_dim.hidden and _column_index(col_key) is not None
                }
                meta.hidden_columns.discard(None)
                meta.merged_cells = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
                meta.tables = _table_metadata(worksheet)

                if cell_count <= _SECOND_PASS_MAX_CELLS:
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.data_type == "f" or (
                                isinstance(cell.value, str) and cell.value.startswith("=")
                            ):
                                formula = str(cell.value)
                                item = {
                                    "cell": cell.coordinate,
                                    "formula": formula,
                                    "cachedValue": None,
                                }
                                if _is_volatile_formula(formula):
                                    item["warning"] = "volatile formula; cached value is not recalculated by XLSX_EXTRACT"
                                meta.formulas.append(item)
                else:
                    meta.warnings.append("Formula scan skipped because sheet exceeds second-pass cell limit.")
                result[worksheet.title] = meta
            return result
        finally:
            workbook.close()

    def _extract_sheet(self, worksheet: Any, meta: SheetMetadata) -> dict[str, Any]:
        hidden = bool(meta.hidden)
        base: dict[str, Any] = {
            "role": "sheet",
            "name": meta.name,
            "index": meta.index,
            "sheetName": meta.name,
            "sheetIndex": meta.index,
            "hidden": hidden,
            "maxRow": meta.max_row,
            "maxColumn": meta.max_column,
            "usedRange": meta.used_range,
            "cellRange": meta.used_range,
            "rowStart": 1,
            "rowEnd": meta.max_row,
            "columnStart": 1,
            "columnEnd": meta.max_column,
            "tables": [],
            "mergedCells": meta.merged_cells,
            "formulas": meta.formulas,
            "warnings": list(meta.warnings),
        }
        if hidden and not self._include_hidden:
            base["indexable"] = False
            base["warning"] = "hidden sheet skipped by default"
            return base

        read = _read_sheet(worksheet, meta)
        rows = read.rows
        headers = _detect_headers(rows)
        compact_text = ""
        if rows:
            compact_text = _rows_to_text(
                sheet_name=meta.name,
                cell_range=meta.used_range,
                rows=rows[: _SECTION_PREVIEW_ROWS + 1],
                headers=headers,
            )
        chunks = _build_chunks(meta, rows, read.row_numbers, headers)
        tables = self._build_tables(meta, read, headers)
        base.update({
            "indexable": True,
            "compactText": _limit(compact_text, _MAX_UNIT_TEXT),
            "rowCount": len(rows),
            "columnCount": _max_row_width(rows),
            "cells": read.cells,
            "tables": tables,
            "chunks": chunks,
        })
        if len(rows) >= _MAX_ROWS_TO_READ:
            base["warnings"].append("Sheet row extraction stopped at safety row limit.")
        if read.cells_truncated:
            base["warnings"].append("Sheet cell metadata truncated at safety cell limit.")
        return base

    def _build_tables(
        self,
        meta: SheetMetadata,
        read: SheetReadResult,
        headers: list[str] | None,
    ) -> list[dict[str, Any]]:
        tables: list[dict[str, Any]] = []
        for table in meta.tables:
            table_range = table.get("cellRange") or table.get("range") or meta.used_range
            selected = _rows_for_range(read.rows, read.row_numbers, table_range)
            table_headers = _detect_headers(selected) or headers
            markdown = _markdown_table(selected, table_headers)
            text = _rows_to_text(
                sheet_name=meta.name,
                cell_range=table_range,
                rows=selected,
                headers=table_headers,
            )
            tables.append({
                **table,
                "role": "table",
                "sheetName": meta.name,
                "sheetIndex": meta.index,
                "type": "named",
                "markdown": _limit(markdown, _MAX_UNIT_TEXT),
                "text": _limit(text, _MAX_UNIT_TEXT),
            })

        if not tables and _is_small_detectable_table(read.rows, meta):
            table_range = meta.used_range
            detected_headers = headers or _detect_headers(read.rows)
            min_col, min_row, max_col, max_row = _range_parts(table_range)
            tables.append({
                "role": "table",
                "tableIndex": 0,
                "name": "DetectedTable1",
                "tableId": "DetectedTable1",
                "sheetName": meta.name,
                "sheetIndex": meta.index,
                "type": "detected",
                "range": table_range,
                "cellRange": table_range,
                "rowStart": min_row,
                "rowEnd": max_row,
                "columnStart": min_col,
                "columnEnd": max_col,
                "rowCount": max(len(read.rows), 0),
                "columnCount": _max_row_width(read.rows),
                "markdown": _limit(_markdown_table(read.rows, detected_headers), _MAX_UNIT_TEXT),
                "text": _limit(
                    _rows_to_text(
                        sheet_name=meta.name,
                        cell_range=table_range,
                        rows=read.rows,
                        headers=detected_headers,
                    ),
                    _MAX_UNIT_TEXT,
                ),
            })
        return tables


class XlsxExtractCapability(Capability):
    name = "XLSX_EXTRACT"

    def __init__(self, service: XlsxExtractService) -> None:
        self._service = service

    def run(self, input: CapabilityInput) -> CapabilityOutput:
        return self._service.run(input)


def _read_sheet(worksheet: Any, meta: SheetMetadata) -> SheetReadResult:
    rows: list[list[str]] = []
    row_numbers: list[int] = []
    cells: list[dict[str, Any]] = []
    cells_truncated = False
    max_col = min(meta.max_column or _MAX_COLUMNS, _MAX_COLUMNS)
    formulas_by_cell = {item["cell"]: item for item in meta.formulas if item.get("cell")}
    for row in worksheet.iter_rows(max_col=max_col):
        row_index = getattr(row[0], "row", len(rows) + 1) if row else len(rows) + 1
        if row_index in meta.hidden_rows:
            continue
        values: list[str] = []
        row_cells: list[dict[str, Any]] = []
        for cell in row:
            col_idx = getattr(cell, "column", None)
            if isinstance(col_idx, str):
                col_idx = _column_index(col_idx)
            if col_idx in meta.hidden_columns:
                continue
            raw_value = getattr(cell, "value", None)
            value = _display_value(raw_value)
            formula_meta = formulas_by_cell.get(getattr(cell, "coordinate", ""))
            if formula_meta is not None:
                formula_meta["cachedValue"] = value or None
                if not value and formula_meta.get("formula"):
                    value = str(formula_meta["formula"])
            if value or formula_meta is not None:
                cell_record: dict[str, Any] = {
                    "cell": getattr(cell, "coordinate", None),
                    "row": row_index,
                    "column": col_idx,
                    "value": value,
                }
                if formula_meta is not None:
                    cell_record["formula"] = formula_meta.get("formula")
                    cell_record["cachedValue"] = formula_meta.get("cachedValue")
                row_cells.append(cell_record)
            values.append(value)
        if any(value for value in values):
            rows.append(_trim_row(values))
            row_numbers.append(row_index)
            if len(cells) < _MAX_CELLS_TO_EXPORT:
                remaining = _MAX_CELLS_TO_EXPORT - len(cells)
                cells.extend(row_cells[:remaining])
                cells_truncated = cells_truncated or len(row_cells) > remaining
            elif row_cells:
                cells_truncated = True
        if len(rows) >= _MAX_ROWS_TO_READ:
            break
    return SheetReadResult(rows, row_numbers, cells, cells_truncated)


def _build_chunks(
    meta: SheetMetadata,
    rows: list[list[str]],
    row_numbers: list[int],
    headers: list[str] | None,
) -> list[dict[str, Any]]:
    if len(rows) <= _CHUNK_DATA_ROWS + 1:
        return []
    chunks: list[dict[str, Any]] = []
    header_index = _header_index(rows, headers) if headers else -1
    header_offset = header_index + 1 if headers else 0
    data_rows = rows[header_offset:]
    data_row_numbers = row_numbers[header_offset:]
    for index in range(0, len(data_rows), _CHUNK_DATA_ROWS):
        window = data_rows[index:index + _CHUNK_DATA_ROWS]
        window_row_numbers = data_row_numbers[index:index + len(window)]
        display_rows = ([rows[header_index]] if headers and header_index >= 0 else []) + window
        row_start = window_row_numbers[0] if window_row_numbers else header_offset + index + 1
        row_end = window_row_numbers[-1] if window_row_numbers else header_offset + index + len(window)
        end_col = max(1, min(_max_row_width(display_rows), _MAX_COLUMNS))
        cell_range = f"A{row_start}:{get_column_letter(end_col)}{row_end}"
        text = _rows_to_text(
            sheet_name=meta.name,
            cell_range=cell_range,
            rows=display_rows,
            headers=headers,
        )
        chunks.append({
            "role": "chunk",
            "chunkIndex": len(chunks),
            "sheetName": meta.name,
            "sheetIndex": meta.index,
            "range": cell_range,
            "cellRange": cell_range,
            "rowStart": row_start,
            "rowEnd": row_end,
            "columnStart": 1,
            "columnEnd": end_col,
            "text": _limit(text, _MAX_UNIT_TEXT),
        })
    return chunks


def _table_metadata(worksheet: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    tables = getattr(worksheet, "tables", None)
    if not tables:
        return result
    values = tables.values() if hasattr(tables, "values") else tables
    for index, table in enumerate(values):
        ref = getattr(table, "ref", None)
        if not ref:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        table_name = getattr(table, "displayName", None) or getattr(table, "name", None)
        result.append({
            "role": "table",
            "tableIndex": index,
            "name": table_name,
            "tableId": table_name or f"Table{index + 1}",
            "range": ref,
            "cellRange": ref,
            "rowStart": min_row,
            "rowEnd": max_row,
            "columnStart": min_col,
            "columnEnd": max_col,
            "rowCount": max_row - min_row + 1,
            "columnCount": max_col - min_col + 1,
        })
    return result


def _rows_for_range(rows: list[list[str]], row_numbers: list[int], cell_range: str) -> list[list[str]]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return rows
    selected: list[list[str]] = []
    for source_row, row_number in zip(rows, row_numbers):
        if min_row <= row_number <= max_row:
            selected.append(source_row[min_col - 1:max_col])
    return [row for row in selected if any(row)]


def _detect_headers(rows: list[list[str]]) -> list[str] | None:
    for first in rows[:10]:
        nonblank = [cell for cell in first if cell]
        if len(nonblank) >= 2:
            break
    else:
        return None
    nonblank = [cell for cell in first if cell]
    if len(nonblank) < 2:
        return None
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(first, start=1):
        header = value.strip() if value else get_column_letter(idx)
        count = seen.get(header, 0)
        seen[header] = count + 1
        headers.append(header if count == 0 else f"{header}_{count + 1}")
    return headers


def _rows_to_text(
    *,
    sheet_name: str,
    cell_range: str,
    rows: list[list[str]],
    headers: list[str] | None,
) -> str:
    lines = [f"[Sheet: {sheet_name}]", f"[Range: {cell_range}]"]
    if not rows:
        return "\n".join(lines)
    header_index = _header_index(rows, headers) if headers else -1
    data_rows = rows[header_index + 1:] if headers else rows
    for row in data_rows:
        parts: list[str] = []
        for idx, value in enumerate(row):
            if not value:
                continue
            header = headers[idx] if headers and idx < len(headers) else get_column_letter(idx + 1)
            parts.append(f"{header}: {value}")
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines).strip()


def _markdown_table(rows: list[list[str]], headers: list[str] | None) -> str:
    if not rows:
        return ""
    table_headers = headers or [get_column_letter(idx + 1) for idx in range(_max_row_width(rows))]
    header_index = _header_index(rows, headers) if headers else -1
    data_rows = rows[header_index + 1:] if headers else rows
    width = len(table_headers)
    lines = [
        "| " + " | ".join(_escape_md(cell) for cell in table_headers) + " |",
        "| " + " | ".join("---" for _ in table_headers) + " |",
    ]
    for row in data_rows[:200]:
        padded = row[:width] + [""] * max(0, width - len(row))
        lines.append("| " + " | ".join(_escape_md(cell) for cell in padded) + " |")
    return "\n".join(lines)


def _header_index(rows: list[list[str]], headers: list[str] | None) -> int:
    if not headers:
        return -1
    for index, row in enumerate(rows):
        normalized = [cell.strip() if cell else "" for cell in row[:len(headers)]]
        if normalized == headers[:len(normalized)]:
            return index
    return 0


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return f"{value:,.6g}"
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return _limit(text, _MAX_CELL_TEXT)


def _is_small_detectable_table(rows: list[list[str]], meta: SheetMetadata) -> bool:
    if not rows or not _detect_headers(rows):
        return False
    width = _max_row_width(rows)
    return len(rows) * max(width, 1) <= _SMALL_TABLE_MAX_CELLS and len(rows) > 1


def _used_range(max_row: int, max_column: int) -> str:
    if max_row <= 0 or max_column <= 0:
        return "A1:A1"
    return f"A1:{get_column_letter(max_column)}{max_row}"


def _range_parts(cell_range: str) -> tuple[int, int, int, int]:
    try:
        return range_boundaries(cell_range)
    except ValueError:
        return (1, 1, 1, 1)


def _max_row_width(rows: list[list[str]]) -> int:
    return max((len(row) for row in rows), default=0)


def _trim_row(values: list[str]) -> list[str]:
    end = len(values)
    while end > 0 and not values[end - 1]:
        end -= 1
    return values[:end]


def _column_index(key: Any) -> Optional[int]:
    if key is None:
        return None
    if isinstance(key, int):
        return key
    text = str(key).strip()
    if not text:
        return None
    total = 0
    for char in text.upper():
        if char < "A" or char > "Z":
            return None
        total = total * 26 + (ord(char) - ord("A") + 1)
    return total


def _is_volatile_formula(formula: str) -> bool:
    upper = formula.upper()
    return any(marker in upper for marker in _VOLATILE_FUNCTIONS)


def _escape_md(value: str) -> str:
    return value.replace("|", "\\|")


def _limit(value: str, max_chars: int) -> str:
    if len(value) <= max_chars:
        return value
    return value[: max_chars - 20].rstrip() + "\n[truncated]"
