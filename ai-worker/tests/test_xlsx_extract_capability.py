"""Tests for the XLSX_EXTRACT capability."""

from __future__ import annotations

import json
from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.capabilities.base import (
    CapabilityError,
    CapabilityInput,
    CapabilityInputArtifact,
)
from app.capabilities.xlsx.artifact_builder import (
    XLSX_MARKDOWN,
    XLSX_PIPELINE_VERSION,
    XLSX_TABLE_JSON,
    XLSX_WORKBOOK_JSON,
)
from app.capabilities.xlsx.service import XlsxExtractCapability, XlsxExtractService


def test_xlsx_extract_emits_workbook_markdown_and_table_artifacts():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_workbook_bytes()))

    assert [artifact.type for artifact in result.outputs] == [
        XLSX_WORKBOOK_JSON,
        XLSX_MARKDOWN,
        XLSX_TABLE_JSON,
    ]
    body = json.loads(result.outputs[0].content)
    assert body["fileType"] == "xlsx"
    assert body["pipelineVersion"] == XLSX_PIPELINE_VERSION
    assert body["sourceRecordId"] == "source-file-1"
    assert body["workbook"]["sheetCount"] == 3

    sheets = {sheet["name"]: sheet for sheet in body["workbook"]["sheets"]}
    assert "매출" in sheets
    assert "요약" in sheets
    assert sheets["숨김"]["hidden"] is True
    assert sheets["숨김"]["indexable"] is False

    sales = sheets["매출"]
    assert sales["role"] == "sheet"
    assert sales["sheetName"] == "매출"
    assert sales["sheetIndex"] == 0
    assert sales["cellRange"].startswith("A1:")
    assert sales["rowStart"] == 1
    assert sales["columnStart"] == 1
    assert sales["usedRange"].startswith("A1:")
    assert sales["mergedCells"] == ["A1:D1"]
    assert any(cell["cell"] == "A1" and cell["row"] == 1 for cell in sales["cells"])
    assert sales["tables"][0]["name"] == "SalesTable"
    assert sales["tables"][0]["role"] == "table"
    assert sales["tables"][0]["sheetName"] == "매출"
    assert sales["tables"][0]["sheetIndex"] == 0
    assert sales["tables"][0]["cellRange"] == "A3:D5"
    assert sales["tables"][0]["rowStart"] == 3
    assert sales["tables"][0]["columnEnd"] == 4
    assert sales["tables"][0]["range"] == "A3:D5"
    assert "직원명: 홍길동" in sales["compactText"]
    assert "매출: 12,000,000" in sales["compactText"]

    summary = sheets["요약"]
    assert summary["formulas"][0]["cell"] == "B2"
    assert summary["formulas"][0]["formula"] == "=SUM(매출!D4:D5)"
    assert summary["formulas"][0]["cachedValue"] is None

    markdown = result.outputs[1].content.decode("utf-8")
    assert "## 매출" in markdown
    assert "SalesTable" in markdown
    assert "숨김" not in markdown

    table_json = json.loads(result.outputs[2].content)
    assert table_json["tables"][0]["sheetName"] == "매출"


def test_xlsx_extract_creates_row_window_chunks_for_large_sheets():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_large_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    assert sheet["name"] == "대용량"
    assert len(sheet["chunks"]) >= 2
    assert sheet["chunks"][0]["role"] == "chunk"
    assert sheet["chunks"][0]["chunkIndex"] == 0
    assert sheet["chunks"][0]["cellRange"].startswith("A2:")
    assert sheet["chunks"][0]["range"].startswith("A2:")
    assert "항목: item-1" in sheet["chunks"][0]["text"]
    assert "값: 1" in sheet["chunks"][0]["text"]
    assert "항목: item-60" in sheet["chunks"][1]["text"]
    assert "값: 60" in sheet["chunks"][1]["text"]


def test_xlsx_extract_keeps_blank_sheet_text_blank_for_indexing_skip():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_blank_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    assert sheet["name"] == "Blank"
    assert sheet["compactText"] == ""
    assert sheet["tables"] == []
    assert sheet["chunks"] == []


def test_xlsx_extract_excludes_hidden_column_formula_values():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_hidden_formula_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    payload = json.dumps(sheet, ensure_ascii=False)
    assert "SECRET123" not in payload
    assert sheet["hiddenColumns"] == ["B"]
    assert sheet["hiddenColumnIndexes"] == [2]
    assert sheet["formulas"] == []
    assert all(cell["column"] != 2 for cell in sheet["cells"])


def test_xlsx_extract_preserves_visible_coordinates_after_hidden_middle_column():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_hidden_middle_column_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    payload = json.dumps(sheet, ensure_ascii=False)
    assert "SECRET123" not in payload
    assert "hidden header" not in payload
    assert sheet["hiddenColumnIndexes"] == [2]
    assert any(
        cell["cell"] == "C2" and cell["column"] == 3 and cell["value"] == "VISIBLE-C"
        for cell in sheet["cells"]
    )
    assert "public_after_hidden: VISIBLE-C" in sheet["compactText"]
    assert "B: VISIBLE-C" not in sheet["compactText"]


def test_xlsx_extract_excludes_protected_hidden_formula_cells():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_protected_hidden_formula_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    payload = json.dumps(sheet, ensure_ascii=False)
    assert "SECRET123" not in payload
    assert sheet["hiddenCells"] == ["B2"]
    assert sheet["formulas"] == []
    assert all(cell["cell"] != "B2" for cell in sheet["cells"])


def test_xlsx_extract_fails_closed_when_hidden_metadata_is_untrusted(monkeypatch):
    service = XlsxExtractService()
    monkeypatch.setattr(service, "_collect_second_pass_metadata", lambda _content, _warnings: {})
    capability = XlsxExtractCapability(service=service)

    result = capability.run(_input(_hidden_formula_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    payload = json.dumps(body, ensure_ascii=False)
    assert "SECRET123" not in payload
    assert body["plainText"] == ""
    assert sheet["indexable"] is False
    assert "hidden row/column/cell metadata could not be verified" in sheet["warning"]


def test_xlsx_extract_fails_closed_for_large_protected_sheet_with_unscanned_hidden_cells():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(_input(_large_protected_hidden_cell_workbook_bytes()))

    body = json.loads(result.outputs[0].content)
    sheet = body["workbook"]["sheets"][0]
    payload = json.dumps(body, ensure_ascii=False)
    assert "SECRET123" not in payload
    assert body["plainText"] == ""
    assert sheet["indexable"] is False
    assert any("protected hidden cells cannot be verified" in warning for warning in sheet["warnings"])


def test_xlsx_extract_rejects_legacy_xls():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    with pytest.raises(CapabilityError) as raised:
        capability.run(
            _input(
                b"not really xls",
                filename="legacy.xls",
                content_type="application/vnd.ms-excel",
            )
        )

    assert raised.value.code == "UNSUPPORTED_INPUT_TYPE"
    assert "Legacy .xls is not supported" in raised.value.message


def test_xlsx_extract_accepts_xlsx_extension_with_octet_stream():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(
        _input(
            _workbook_bytes(),
            filename="sales.xlsx",
            content_type="application/octet-stream",
        )
    )

    body = json.loads(result.outputs[0].content)
    assert body["fileType"] == "xlsx"


def test_xlsx_extract_allows_xlsm_without_macro_execution():
    capability = XlsxExtractCapability(service=XlsxExtractService())

    result = capability.run(
        _input(
            _workbook_bytes(),
            filename="macro.xlsm",
            content_type="application/vnd.ms-excel.sheet.macroenabled.12",
        )
    )

    body = json.loads(result.outputs[0].content)
    assert body["fileType"] == "xlsm"
    assert body["security"]["macrosExecuted"] is False
    assert any("macros are not executed" in warning for warning in body["warnings"])


def test_registry_registers_xlsx_extract_without_ocr_or_rag():
    from app.capabilities import registry as registry_module
    from app.core.config import WorkerSettings

    settings = WorkerSettings(
        rag_enabled=False,
        ocr_enabled=False,
        ocr_extract_enabled=False,
        multimodal_enabled=False,
        xlsx_extract_enabled=True,
        pdf_extract_enabled=False,
    )

    result = registry_module.build_default_registry(settings)

    assert result.available() == ["MOCK", "XLSX_EXTRACT"]


def _input(
    content: bytes,
    *,
    filename: str = "sales.xlsx",
    content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> CapabilityInput:
    return CapabilityInput(
        job_id="job-1",
        capability="XLSX_EXTRACT",
        attempt_no=1,
        inputs=[
            CapabilityInputArtifact(
                artifact_id="input-artifact-1",
                source_file_id="source-file-1",
                type="INPUT_FILE",
                content=content,
                content_type=content_type,
                filename=filename,
            )
        ],
    )


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    ws.merge_cells("A1:D1")
    ws["A1"] = "2024 매출"
    ws.append([])
    ws.append(["직원명", "연도", "지역", "매출"])
    ws.append(["홍길동", 2024, "서울", 12_000_000])
    ws.append(["김철수", 2024, "부산", 9_000_000])
    table = Table(displayName="SalesTable", ref="A3:D5")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    summary = wb.create_sheet("요약")
    summary.append(["항목", "값"])
    summary.append(["총매출", "=SUM(매출!D4:D5)"])

    hidden = wb.create_sheet("숨김")
    hidden.sheet_state = "hidden"
    hidden.append(["secret", "value"])

    return _save(wb)


def _large_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "대용량"
    ws.append(["항목", "값"])
    for index in range(1, 130):
        ws.append([f"item-{index}", index])
    return _save(wb)


def _blank_workbook_bytes() -> bytes:
    wb = Workbook()
    wb.active.title = "Blank"
    return _save(wb)


def _hidden_formula_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Visible"
    ws.append(["name", "secret"])
    ws.append(["public", "not-secret"])
    ws.column_dimensions["B"].hidden = True
    ws["B2"] = '="SECRET123"'
    return _save(wb)


def _hidden_middle_column_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Visible"
    ws.append(["name", "hidden header", "public_after_hidden"])
    ws.append(["row-1", "SECRET123", "VISIBLE-C"])
    ws.column_dimensions["B"].hidden = True
    return _save(wb)


def _protected_hidden_formula_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Visible"
    ws.append(["name", "protected_formula"])
    ws.append(["public", None])
    ws["B2"] = '="SECRET123"'
    ws["B2"].protection = Protection(hidden=True, locked=True)
    ws.protection.sheet = True
    return _save(wb)


def _large_protected_hidden_cell_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Protected"
    ws.append(["name", "protected_formula"])
    ws.append(["public", None])
    ws["B2"] = '="SECRET123"'
    ws["B2"].protection = Protection(hidden=True, locked=True)
    ws["CV501"] = "size-marker"
    ws.protection.sheet = True
    return _save(wb)


def _save(wb: Workbook) -> bytes:
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
