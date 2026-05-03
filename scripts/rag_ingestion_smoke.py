"""Smoke test for xlsx RAG ingestion v2.

Preconditions:
  - PostgreSQL and Redis are running.
  - core-api is running on the selected base URL.
  - ai-worker is running with XLSX_EXTRACT enabled.
  - If DB validation is enabled, Docker can exec into the PostgreSQL container.

The script uploads a generated XLSX file, starts XLSX_EXTRACT, waits for the
job to finish, and checks that v2 document/search-unit rows were created with
parser version, location JSON, and citation text.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_BASE_URL = "http://localhost:8080"
DEFAULT_DB_CONTAINER = "aipipeline-postgres"
DEFAULT_DB_USER = "aipipeline"
DEFAULT_DB_NAME = "aipipeline"
DEFAULT_REPORT = Path("reports/rag_ingestion_smoke_report.json")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    run_report: dict[str, Any] = {
        "run_id": utc_run_id(),
        "status": "FAILED",
        "source_file_id": None,
        "job_id": None,
        "job_status": None,
        "db_report": None,
        "warnings": [],
    }
    with tempfile.TemporaryDirectory(prefix="rag-ingestion-smoke-") as tmp:
        xlsx_path = Path(tmp) / "rag-ingestion-sales.xlsx"
        write_sample_xlsx(xlsx_path)

        client = httpx.Client(base_url=args.base_url, timeout=args.http_timeout)
        source = upload_source(client, xlsx_path)
        source_id = source["sourceFileId"]
        run_report["source_file_id"] = source_id
        print(f"[1/4] uploaded sourceFileId={source_id} name={source['originalFileName']}")

        job = start_xlsx_extract(client, source_id)
        job_id = job["jobId"]
        run_report["job_id"] = job_id
        print(f"[2/4] started XLSX_EXTRACT jobId={job_id}")

        final = wait_job(client, job_id, timeout_seconds=args.poll_timeout)
        run_report["job_status"] = final["status"]
        print(f"[3/4] job status={final['status']}")
        if final["status"] != "SUCCEEDED":
            run_report["job"] = final
            write_json_report(Path(args.report), run_report)
            print(json.dumps(final, ensure_ascii=False, indent=2))
            return 3

        if not args.skip_db_check:
            db_report = query_ingestion_report(
                container=args.db_container,
                user=args.db_user,
                database=args.db_name,
                source_file_id=source_id,
            )
            validate_report(db_report)
            run_report["db_report"] = db_report
            print("[4/4] DB v2 rows verified")
            print(json.dumps(db_report, ensure_ascii=False, indent=2))
        else:
            run_report["warnings"].append("DB validation skipped")
            print("[4/4] DB check skipped")

    run_report["status"] = "PASSED"
    write_json_report(Path(args.report), run_report)
    return 0


def parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--http-timeout", type=float, default=20.0)
    parser.add_argument("--poll-timeout", type=float, default=60.0)
    parser.add_argument("--db-container", default=DEFAULT_DB_CONTAINER)
    parser.add_argument("--db-user", default=DEFAULT_DB_USER)
    parser.add_argument("--db-name", default=DEFAULT_DB_NAME)
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument("--skip-db-check", action="store_true")
    return parser.parse_args(argv)


def write_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출현황"
    ws.merge_cells("A1:F1")
    ws["A1"] = "2024년 분기별 매출"
    ws.append([])
    ws.append(["기간", "부서", "제품군", "매출", "전년매출", "증감률"])
    ws.append(["2024 Q3", "온라인", "가전", 12_000_000_000, 11_100_000_000, "=D4/E4-1"])
    ws.append(["2024 Q3", "오프라인", "생활", 9_000_000_000, 8_700_000_000, "=D5/E5-1"])
    table = Table(displayName="SalesQuarterTable", ref="A3:F5")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    ws.column_dimensions["E"].hidden = True

    hidden = wb.create_sheet("숨김")
    hidden.sheet_state = "hidden"
    hidden.append(["secret", "value"])

    wb.save(path)


def upload_source(client: httpx.Client, path: Path) -> dict[str, Any]:
    with path.open("rb") as handle:
        response = client.post(
            "/api/v1/library/source-files",
            files={
                "file": (
                    path.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    response.raise_for_status()
    return response.json()


def start_xlsx_extract(client: httpx.Client, source_file_id: str) -> dict[str, Any]:
    response = client.post(f"/api/v1/library/source-files/{source_file_id}/xlsx-extract")
    response.raise_for_status()
    return response.json()


def wait_job(client: httpx.Client, job_id: str, *, timeout_seconds: float) -> dict[str, Any]:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        response = client.get(f"/api/v1/jobs/{job_id}")
        response.raise_for_status()
        body = response.json()
        if body["status"] in {"SUCCEEDED", "FAILED", "CANCELLED"}:
            return body
        time.sleep(0.5)
    raise TimeoutError(f"Timed out waiting for job {job_id}")


def query_ingestion_report(
    *,
    container: str,
    user: str,
    database: str,
    source_file_id: str,
) -> dict[str, Any]:
    sql = f"""
    select jsonb_build_object(
      'source_file_id', {sql_literal(source_file_id)},
      'source_status', (select status from source_file where id = {sql_literal(source_file_id)}),
      'extracted_artifact_count', (
        select count(*) from extracted_artifact where source_file_id = {sql_literal(source_file_id)}
      ),
      'document_version_count', (
        select count(*) from document_version where source_file_id = {sql_literal(source_file_id)}
      ),
      'parsed_artifact_count', (
        select count(*) from parsed_artifact where source_file_id = {sql_literal(source_file_id)}
      ),
      'search_unit_count', (
        select count(*) from search_unit where source_file_id = {sql_literal(source_file_id)}
      ),
      'v2_ready_search_unit_count', (
        select count(*) from search_unit
        where source_file_id = {sql_literal(source_file_id)}
          and parser_version is not null
          and location_json is not null
          and citation_text is not null
      ),
      'missing_citation_count', (
        select count(*) from search_unit
        where source_file_id = {sql_literal(source_file_id)}
          and source_file_type in ('SPREADSHEET', 'PDF')
          and (parser_version is null or location_json is null or citation_text is null)
      ),
      'table_metadata_count', (
        select count(*) from table_metadata where source_file_id = {sql_literal(source_file_id)}
      ),
      'cell_metadata_count', (
        select count(*) from cell_metadata where source_file_id = {sql_literal(source_file_id)}
      ),
      'formula_cell_metadata_count', (
        select count(*) from cell_metadata
        where source_file_id = {sql_literal(source_file_id)}
          and formula is not null
      ),
      'hidden_search_unit_leakage_count', (
        select count(*) from search_unit
        where source_file_id = {sql_literal(source_file_id)}
          and lower(source_file_type) in ('spreadsheet', 'xlsx')
          and (
            location_json->>'hidden' = 'true'
            or location_json->>'hidden_sheet' = 'true'
            or coalesce(citation_text, '') like '%숨김%'
          )
      ),
      'samples', (
        select coalesce(jsonb_agg(jsonb_build_object(
          'unit_type', unit_type,
          'chunk_type', chunk_type,
          'location_type', location_type,
          'citation_text', citation_text,
          'parser_version', parser_version,
          'location_json', location_json
        ) order by unit_type, unit_key), '[]'::jsonb)
        from (
          select * from search_unit
          where source_file_id = {sql_literal(source_file_id)}
          order by unit_type, unit_key
          limit 5
        ) sample
      )
    );
    """
    completed = subprocess.run(
        [
            "docker",
            "exec",
            container,
            "psql",
            "-U",
            user,
            "-d",
            database,
            "-At",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return json.loads(completed.stdout.strip())


def validate_report(report: dict[str, Any]) -> None:
    checks = {
        "source_status": report.get("source_status") == "READY",
        "extracted_artifact_count": int(report.get("extracted_artifact_count") or 0) >= 3,
        "document_version_count": int(report.get("document_version_count") or 0) >= 1,
        "parsed_artifact_count": int(report.get("parsed_artifact_count") or 0) >= 1,
        "search_unit_count": int(report.get("search_unit_count") or 0) >= 3,
        "v2_ready_search_unit_count": int(report.get("v2_ready_search_unit_count") or 0) >= 3,
        "missing_citation_count": int(report.get("missing_citation_count") or 0) == 0,
        "table_metadata_count": int(report.get("table_metadata_count") or 0) > 0,
        "cell_metadata_count": int(report.get("cell_metadata_count") or 0) > 0,
        "hidden_search_unit_leakage_count": int(report.get("hidden_search_unit_leakage_count") or 0) == 0,
    }
    failed = [name for name, ok in checks.items() if not ok]
    if failed:
        raise AssertionError(f"RAG ingestion smoke validation failed: {failed}")


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def write_json_report(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def utc_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%SZ")


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"[FAIL] {exc}", file=sys.stderr)
        sys.exit(1)
